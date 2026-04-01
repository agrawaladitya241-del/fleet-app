import streamlit as st
import pandas as pd
from datetime import date
import openpyxl
import re

from driver_helper import (
    process_driver_file,
    driver_summary,
    driver_home_days
)
from database import save_fleet_data, save_driver_data


# ================= GET LATEST FILE =================
def get_latest_file(files):

    dated_files = []

    for f in files:
        name = f.name

        match = re.search(r'(\d{2})\.(\d{2})', name)
        if match:
            day, month = match.groups()
            date_key = int(month) * 100 + int(day)
            dated_files.append((date_key, f))

    if not dated_files:
        return files[-1]

    return sorted(dated_files, reverse=True)[0][1]


# ================= FLEET PROCESSING =================
def process_file(uploaded_file):
    wb = openpyxl.load_workbook(uploaded_file)
    ws = wb.active

    vehicle_col = None
    trips_col = None

    for row in ws.iter_rows(min_row=1, max_row=10):
        for i, cell in enumerate(row):
            if cell.value:
                text = str(cell.value).upper()
                if "VEHICLE" in text:
                    vehicle_col = i
                if "TRIP" in text:
                    trips_col = i

    data = {}

    if vehicle_col is None or trips_col is None:
        return data

    for row in ws.iter_rows(min_row=2):

        vehicle_raw = row[vehicle_col].value
        if not vehicle_raw:
            continue

        vehicle = str(vehicle_raw).replace(" ", "").upper()

        try:
            trips = float(row[trips_col].value)
        except:
            trips = 0

        idle = 0
        for i, cell in enumerate(row):
            if i in [vehicle_col, trips_col]:
                continue

            val = str(cell.value).upper() if cell.value else ""
            if "WAIT" in val or "PARK" in val:
                idle += 1

        if vehicle not in data:
            data[vehicle] = {"trips": 0, "idle": 0}

        data[vehicle]["trips"] += trips
        data[vehicle]["idle"] += idle

    return data


def fleet_summary(files):
    data = process_file(files[0])

    total_vehicles = len(data)
    total_trips = sum(v["trips"] for v in data.values())
    total_idle = sum(v["idle"] for v in data.values())

    efficiency = round(total_trips / (total_trips + total_idle), 3) if (total_trips + total_idle) else 0

    return {
        "total_vehicles": total_vehicles,
        "total_trips": int(total_trips),
        "total_idle": total_idle,
        "efficiency": efficiency,
        "vehicle_data": data
    }


# ================= STATUS =================
def extract_fleet_status(file):

    final = {}
    df = pd.read_excel(file)

    for _, row in df.iterrows():

        vehicle = None
        for col in df.columns:
            if "vehicle" in str(col).lower():
                vehicle = str(row[col]).strip().upper()
                break

        if not vehicle or vehicle == "NAN":
            continue

        row_values = [str(x).upper() for x in row if pd.notna(x)]

        dh = sum(1 for x in row_values if "DH" in x)
        dp = sum(1 for x in row_values if "DP" in x)

        current_status = ""
        for val in reversed(row_values):
            if val.strip():
                current_status = val
                break

        if "DH" in current_status:
            status_type = "Driver Home"
        elif "DP" in current_status:
            status_type = "Delay"
        else:
            status_type = "Active"

        final[vehicle] = {
            "DH": dh,
            "DP": dp,
            "status_type": status_type
        }

    return pd.DataFrame(final).T.reset_index().rename(columns={"index": "vehicle"})


# ================= MONTHLY ANALYSIS =================
def monthly_analysis(file):

    df = pd.read_excel(file)

    vehicle_col = None
    for col in df.columns:
        if "vehicle" in str(col).lower():
            vehicle_col = col
            break

    if vehicle_col is None:
        return None

    results = []

    for _, row in df.iterrows():

        vehicle = str(row[vehicle_col]).strip().upper()

        if not vehicle or vehicle == "NAN":
            continue

        dp = 0
        dh = 0
        trips = 0
        days = 0

        for val in row:

            if pd.isna(val):
                continue

            val_str = str(val).upper()
            days += 1

            if "DP" in val_str:
                dp += 1
            elif "DH" in val_str:
                dh += 1

            try:
                num = float(val)
                if num > 0:
                    trips += num
            except:
                pass

        results.append({
            "vehicle": vehicle,
            "DP": dp,
            "DH": dh,
            "Trips": trips,
            "Days": days,
            "Avg_DP": round(dp / days, 2) if days else 0,
            "Avg_DH": round(dh / days, 2) if days else 0,
            "Avg_Trips": round(trips / days, 2) if days else 0
        })

    return pd.DataFrame(results)


# ================= SMART QUERY =================
def smart_query(query, df):

    query = query.lower()

    if "dp" in query:
        return df.sort_values(by="DP", ascending=False).head(10)

    if "dh" in query:
        return df.sort_values(by="DH", ascending=False).head(10)

    if "trip" in query:
        return df.sort_values(by="Trips", ascending=False).head(10)

    if "active" in query:
        return df[df["status_type"] == "Active"]

    return "Query not understood"


# ================= APP =================
st.set_page_config(page_title="Fleet Intelligence System", layout="wide")
st.title("🚛 Fleet Intelligence System")

tab1, tab2 = st.tabs(["Fleet", "Driver"])


# ================= FLEET =================
with tab1:

    files = st.file_uploader("Upload Fleet Files", type=["xlsx"], accept_multiple_files=True)

    if files:

        latest_file = get_latest_file(files)

        summary = fleet_summary([latest_file])
        status_df = extract_fleet_status(latest_file)
        monthly_df = monthly_analysis(latest_file)

        save_fleet_data(summary["vehicle_data"], str(date.today()))

        # TOP METRICS
        col1, col2, col3, col4 = st.columns(4)
        col1.metric("Vehicles", summary["total_vehicles"])
        col2.metric("Trips", summary["total_trips"])
        col3.metric("Idle", summary["total_idle"])
        col4.metric("Efficiency", summary["efficiency"])

        # STATUS
        st.subheader("🚛 Fleet Status Dashboard")

        c1, c2, c3 = st.columns(3)
        c1.metric("Active", len(status_df[status_df["status_type"] == "Active"]))
        c2.metric("Driver Home", len(status_df[status_df["status_type"] == "Driver Home"]))
        c3.metric("Delayed", len(status_df[status_df["status_type"] == "Delay"]))

        st.dataframe(status_df)

        # MONTHLY INTELLIGENCE
        st.divider()
        st.subheader("📊 Monthly Intelligence")

        if monthly_df is not None:

            m1, m2, m3 = st.columns(3)
            m1.metric("Avg DP", round(monthly_df["DP"].mean(), 2))
            m2.metric("Avg DH", round(monthly_df["DH"].mean(), 2))
            m3.metric("Avg Trips", round(monthly_df["Trips"].mean(), 2))

            st.subheader("🚨 Top Delay Vehicles")
            st.dataframe(monthly_df.sort_values(by="DP", ascending=False).head(10))

            st.subheader("🏠 Driver Issues")
            st.dataframe(monthly_df.sort_values(by="DH", ascending=False).head(10))

            st.subheader("🚛 Best Vehicles")
            st.dataframe(monthly_df.sort_values(by="Trips", ascending=False).head(10))

        # SMART QUERY
        st.divider()
        st.subheader("🔍 Smart Query")

        query = st.text_input("Ask (e.g., top dp, top trips)")

        if st.button("Run Query"):

            result = smart_query(query, monthly_df)

            if isinstance(result, str):
                st.warning(result)
            else:
                st.dataframe(result)


# ================= DRIVER =================
with tab2:

    file = st.file_uploader("Upload Driver File")

    if file:
        df = process_driver_file(file)

        save_driver_data(df)

        st.subheader("Driver Summary")
        st.dataframe(driver_summary(df))

        st.subheader("Driver Home Days")
        st.dataframe(driver_home_days(df))
