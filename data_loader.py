"""
data_loader.py
--------------
Reads the multi-sheet fleet Excel file into tidy long-format DataFrames.

Handles the messy reality of the source file:
  - Multiple month sheets (February, March, APRIL) plus junk sheets to ignore.
  - Date column headers in any of 4 formats:
      1. Real datetime cells
      2. Text like '01-Feb', '01.APRL', '5-Aril', '03-Aprl', '6 aprl'
      3. Pure day numbers
  - Plant-In/Out columns interleaved between date columns with messy content:
      - Timestamps: 'in-16.03/11 AM\nout-17.03/12 AM', 'IN-24.03/04PM out-26.03/03PM'
      - Service notes: 'Clutch Adjust & silencer Leak', 'TYRE REQUIR'
      - Location flags: 'L POINT', 'T DUBURI'
  - Hidden rows and hidden columns (openpyxl reads these by default — we keep that).
  - Optional 'Trip' column at the end (March has it, April doesn't).
  - Yellow-highlighted cells marking routes for analysis.
  - Accident vehicles with 'ACCIDENT' / 'Accidental Work' tags.

Output: two long-format DataFrames:
  - daily_df:  one row per (vehicle, day) — status_raw + is_yellow + manual_trip_count
  - plant_df:  one row per plant in/out event — parsed in_time, out_time, dwell_hours
"""

from __future__ import annotations

import re
from datetime import datetime, time
from typing import Dict, List, Optional, Tuple

import pandas as pd
from openpyxl import load_workbook
from openpyxl.utils import get_column_letter


# ------------------------------------------------------------------
# Month recognition
# ------------------------------------------------------------------

_MONTH_PATTERNS = {
    "jan": (1, "January"),
    "feb": (2, "February"),
    "mar": (3, "March"),
    "apr": (4, "April"),
    "may": (5, "May"),
    "jun": (6, "June"),
    "jul": (7, "July"),
    "aug": (8, "August"),
    "sep": (9, "September"),
    "oct": (10, "October"),
    "nov": (11, "November"),
    "dec": (12, "December"),
}


def _identify_month(sheet_name: str) -> Optional[Tuple[int, str]]:
    lower = sheet_name.strip().lower()
    for key, (num, name) in _MONTH_PATTERNS.items():
        if lower.startswith(key) or lower == key:
            return (num, name)
    return None


def list_month_sheets(file_path_or_buffer) -> List[Tuple[str, int, str]]:
    """Return (sheet_name, month_number, display_name) for each month sheet."""
    wb = load_workbook(file_path_or_buffer, read_only=False, data_only=True)
    found = []
    for sheet_name in wb.sheetnames:
        match = _identify_month(sheet_name)
        if match:
            found.append((sheet_name, match[0], match[1]))
    wb.close()
    return sorted(found, key=lambda t: t[1])


# ------------------------------------------------------------------
# Header classification
# ------------------------------------------------------------------

_META_KEYWORDS = [
    "sl no", "sl.no", "serial",
    "vehicle", "truck",
    "model",
    "driver",
    "cont no", "contact", "phone",
    "total trip", "trip count",
    "location", "time in", "time out",
    "running km",
]

_PLANT_KEYWORDS = ["plant in", "plant out", "plant/in", "plantin", "plnat in", "pplantin"]

# Matches text-date headers like '01-Feb', '01.APRL', '5-Aril', '6 aprl'
_DATE_TEXT_RE = re.compile(
    r"^\s*(\d{1,2})[\s\-\.\/]\s*"
    r"(jan|feb|mar|apr|apri|aprl|april|ari|aril|aprl|may|jun|jul|aug|sep|oct|nov|dec)",
    re.IGNORECASE,
)


def _is_meta_header(value) -> bool:
    if value is None or not isinstance(value, str):
        return False
    lower = value.strip().lower()
    return any(kw in lower for kw in _META_KEYWORDS)


def _is_plant_header(value) -> bool:
    if value is None or not isinstance(value, str):
        return False
    lower = value.strip().lower().replace(" ", "").replace("/", "")
    return any(kw.replace(" ", "").replace("/", "") in lower for kw in _PLANT_KEYWORDS)


def _is_trip_column_header(value) -> bool:
    """Detect the manual 'Trip' count column (March has this)."""
    if value is None or not isinstance(value, str):
        return False
    v = value.strip().lower()
    return v == "trip" or v == "trips" or "total trip" in v


def _is_status_column_header(value) -> bool:
    """Detect the 'status' column at end of sheets (not a date)."""
    if value is None or not isinstance(value, str):
        return False
    return value.strip().lower() == "status"


def _parse_date_header(value, fallback_year: int, fallback_month: int) -> Optional[datetime]:
    """Turn a header cell into a datetime if possible, else None."""
    if isinstance(value, datetime):
        return value
    if value is None:
        return None

    s = str(value).strip()
    if not s:
        return None

    # Pure day number: "1", "2", "31"
    if re.match(r"^\d{1,2}$", s):
        day = int(s)
        if 1 <= day <= 31:
            try:
                return datetime(fallback_year, fallback_month, day)
            except ValueError:
                return None

    # Text date like '01-Feb', '05.APRL', '5-Aril', '6 aprl'
    m = _DATE_TEXT_RE.match(s)
    if m:
        day = int(m.group(1))
        month_match = re.search(
            r"(jan|feb|mar|apr|apri|aprl|april|ari|aril|may|jun|jul|aug|sep|oct|nov|dec)",
            s, re.IGNORECASE,
        )
        if month_match:
            abbr = month_match.group(1).lower()[:3]
            if abbr in ("ari", "apr"):
                abbr = "apr"
            for key, (num, _) in _MONTH_PATTERNS.items():
                if key == abbr:
                    try:
                        return datetime(fallback_year, num, day)
                    except ValueError:
                        return None
        try:
            return datetime(fallback_year, fallback_month, day)
        except ValueError:
            return None

    return None


# ------------------------------------------------------------------
# Fill color detection (for yellow = route)
# ------------------------------------------------------------------

def _is_yellow(cell) -> bool:
    """Return True if the cell has a yellow-ish fill color."""
    if cell is None or cell.fill is None or cell.fill.fgColor is None:
        return False
    color = cell.fill.fgColor.rgb if cell.fill.fgColor.type == "rgb" else cell.fill.fgColor.value
    if not color or color == "00000000":
        return False
    color_str = str(color).upper()
    # Yellow variants: FFFF00, FFFFFF00, FFFFC000 (amber), etc.
    # We consider anything with 'FF' in red+green bytes and low blue as yellow.
    # Simpler heuristic: contains 'FFFF00' or ends with 'FF00'.
    if "FFFF00" in color_str:
        return True
    # Amber/orange is also used for routes in some sheets
    if "FFC000" in color_str or "FFA500" in color_str:
        return True
    return False


# ------------------------------------------------------------------
# Plant In/Out timestamp parsing
# ------------------------------------------------------------------

# Patterns of "IN" and "OUT" timestamps we see in the data:
#   'in-16.03/11 AM\nout-17.03/12 AM'
#   'IN-24.03/04PM out-26.03/03PM'
#   'in- 12.03(09 PM) Out- 13.03(07 AM)'
#   '07.03(9.00PM)OUT-08.03(2.00PM)'
#   '10.3 (4:18PM) 11.3 (2 AM)'
#   'in-31.03/08 PM'    <-- no OUT

_DT_PART = r"(\d{1,2})\s*[\.\/\-]\s*(\d{1,2})"   # day.month or day/month
_TIME_PART = (
    r"(\d{1,2})\s*[:\.]?\s*(\d{0,2})\s*"
    r"(AM|PM|am|pm|A\.M\.|P\.M\.)?"
)
_IN_RE = re.compile(
    rf"(?:in|IN)\s*[\-:]?\s*{_DT_PART}\s*[\s\/\(]?\s*{_TIME_PART}",
    re.IGNORECASE,
)
_OUT_RE = re.compile(
    rf"(?:out|OUT)\s*[\-:]?\s*{_DT_PART}\s*[\s\/\(]?\s*{_TIME_PART}",
    re.IGNORECASE,
)
# Fallback: "7.03(9PM)" with no "in"/"out" keyword — just a date+time
_BARE_DT_RE = re.compile(rf"{_DT_PART}\s*[\(\s\/]?\s*{_TIME_PART}")


def _build_datetime(
    day: int, month: int, year: int,
    hour: int, minute: int, ampm: Optional[str]
) -> Optional[datetime]:
    try:
        if ampm:
            ampm = ampm.lower().replace(".", "")
            if ampm.startswith("p") and hour < 12:
                hour += 12
            elif ampm.startswith("a") and hour == 12:
                hour = 0
        if hour > 23 or minute > 59:
            return None
        return datetime(year, month, day, hour, minute)
    except (ValueError, TypeError):
        return None


def parse_plant_inout(text: str, fallback_year: int) -> Dict:
    """
    Parse a plant in/out cell. Returns a dict with:
      in_time, out_time, dwell_hours, note
    Any of them may be None.
    """
    result = {"in_time": None, "out_time": None, "dwell_hours": None, "note": None}
    if not text or not isinstance(text, str):
        return result

    # Normalize whitespace
    s = text.strip().replace("\n", " ")

    # If text contains no digits at all, it's a note (like "Clutch Adjust" or "L POINT")
    if not re.search(r"\d", s):
        result["note"] = s
        return result

    in_match = _IN_RE.search(s)
    out_match = _OUT_RE.search(s)

    if in_match:
        d, mo, h, mn, ap = in_match.groups()
        result["in_time"] = _build_datetime(
            int(d), int(mo), fallback_year,
            int(h), int(mn or 0), ap,
        )

    if out_match:
        d, mo, h, mn, ap = out_match.groups()
        result["out_time"] = _build_datetime(
            int(d), int(mo), fallback_year,
            int(h), int(mn or 0), ap,
        )

    # Fallback: if no IN/OUT keywords but two bare date+times, assume first=in, second=out
    if not in_match and not out_match:
        matches = list(_BARE_DT_RE.finditer(s))
        if len(matches) >= 2:
            d1, mo1, h1, mn1, ap1 = matches[0].groups()
            d2, mo2, h2, mn2, ap2 = matches[1].groups()
            result["in_time"] = _build_datetime(
                int(d1), int(mo1), fallback_year, int(h1), int(mn1 or 0), ap1
            )
            result["out_time"] = _build_datetime(
                int(d2), int(mo2), fallback_year, int(h2), int(mn2 or 0), ap2
            )
        elif len(matches) == 1:
            d, mo, h, mn, ap = matches[0].groups()
            result["in_time"] = _build_datetime(
                int(d), int(mo), fallback_year, int(h), int(mn or 0), ap
            )
        else:
            # Couldn't parse anything — stash as note
            result["note"] = s

    # Compute dwell
    if result["in_time"] and result["out_time"]:
        delta = result["out_time"] - result["in_time"]
        hours = delta.total_seconds() / 3600.0
        # Guard against negative or absurd values (parsing errors)
        if 0 <= hours <= 24 * 14:
            result["dwell_hours"] = round(hours, 2)

    return result


# ------------------------------------------------------------------
# Core loader
# ------------------------------------------------------------------

def load_month_sheet(
    file_path_or_buffer,
    sheet_name: str,
    month_number: int,
    year: int,
) -> Tuple[pd.DataFrame, pd.DataFrame]:
    """
    Load one month sheet. Returns two DataFrames:
      daily_df: one row per (vehicle, day)
      plant_df: one row per plant in/out cell (parsed)
    """
    wb = load_workbook(file_path_or_buffer, data_only=True)
    ws = wb[sheet_name]

    header_row_idx = 1
    header_row = [ws.cell(row=header_row_idx, column=c).value for c in range(1, ws.max_column + 1)]

    # Classify every column
    col_info: Dict[int, Dict] = {}
    for col_idx_0, header_val in enumerate(header_row):
        col_idx = col_idx_0 + 1  # 1-based
        if _is_trip_column_header(header_val):
            col_info[col_idx] = {"kind": "trip_count", "name": str(header_val).strip()}
            continue
        if _is_status_column_header(header_val):
            col_info[col_idx] = {"kind": "status_text", "name": str(header_val).strip()}
            continue
        if _is_meta_header(header_val):
            col_info[col_idx] = {"kind": "meta", "name": str(header_val).strip().lower()}
            continue
        if _is_plant_header(header_val):
            col_info[col_idx] = {"kind": "plant", "name": str(header_val).strip()}
            continue
        parsed = _parse_date_header(header_val, year, month_number)
        if parsed is not None:
            col_info[col_idx] = {"kind": "date", "date": parsed}
            continue
        col_info[col_idx] = {"kind": "skip"}

    # Locate key meta cols
    def find_meta_col(keywords: List[str]) -> Optional[int]:
        for idx, info in col_info.items():
            if info["kind"] != "meta":
                continue
            name = info.get("name", "")
            for kw in keywords:
                if kw in name:
                    return idx
        return None

    vehicle_col = find_meta_col(["vehicle", "truck"])
    driver_col = find_meta_col(["driver"])
    contact_col = find_meta_col(["cont no", "contact", "phone"])
    model_col = find_meta_col(["model"])

    # Locate the Trip count column if present
    trip_count_col = next(
        (idx for idx, info in col_info.items() if info["kind"] == "trip_count"),
        None,
    )

    date_cols = [(idx, info["date"]) for idx, info in col_info.items() if info["kind"] == "date"]
    plant_cols = [(idx, info["name"]) for idx, info in col_info.items() if info["kind"] == "plant"]

    if vehicle_col is None:
        wb.close()
        return pd.DataFrame(), pd.DataFrame()

    # For each plant column, find the nearest preceding date column (so we can
    # attribute the plant in/out event to the correct day).
    plant_to_date = {}
    for p_idx, _ in plant_cols:
        nearest_date = None
        for d_idx, d_date in date_cols:
            if d_idx < p_idx:
                nearest_date = d_date
            else:
                break
        plant_to_date[p_idx] = nearest_date

    daily_records = []
    plant_records = []
    month_display = _MONTH_PATTERNS[next(k for k, v in _MONTH_PATTERNS.items() if v[0] == month_number)][1]

    for row_idx in range(header_row_idx + 1, ws.max_row + 1):
        vehicle_raw = ws.cell(row=row_idx, column=vehicle_col).value
        if vehicle_raw is None:
            continue
        vehicle = str(vehicle_raw).strip().upper()
        if len(vehicle) < 3 or vehicle == "FD":
            continue

        driver = ""
        if driver_col is not None:
            dv = ws.cell(row=row_idx, column=driver_col).value
            if dv is not None:
                driver = str(dv).strip()

        contact = ""
        if contact_col is not None:
            cv = ws.cell(row=row_idx, column=contact_col).value
            if cv is not None:
                contact = str(cv).strip()

        model = ""
        if model_col is not None:
            mv = ws.cell(row=row_idx, column=model_col).value
            if mv is not None:
                model = str(mv).strip()

        manual_trip = None
        if trip_count_col is not None:
            tv = ws.cell(row=row_idx, column=trip_count_col).value
            if isinstance(tv, (int, float)):
                manual_trip = int(tv)

        # Daily cells
        for d_idx, d_date in date_cols:
            cell = ws.cell(row=row_idx, column=d_idx)
            val = cell.value
            status_raw = str(val).strip() if val is not None else ""
            is_yellow = _is_yellow(cell) if status_raw else False

            daily_records.append({
                "vehicle": vehicle,
                "driver": driver,
                "contact": contact,
                "model": model,
                "date": d_date,
                "status_raw": status_raw,
                "is_yellow": is_yellow,
                "month_name": month_display,
                "manual_trip_count": manual_trip,
            })

        # Plant in/out cells
        for p_idx, p_name in plant_cols:
            val = ws.cell(row=row_idx, column=p_idx).value
            if val is None:
                continue
            text = str(val).strip()
            if not text:
                continue
            parsed = parse_plant_inout(text, fallback_year=year)
            attributed_date = plant_to_date.get(p_idx)
            plant_records.append({
                "vehicle": vehicle,
                "date": attributed_date,
                "month_name": month_display,
                "raw_text": text,
                "in_time": parsed["in_time"],
                "out_time": parsed["out_time"],
                "dwell_hours": parsed["dwell_hours"],
                "note": parsed["note"],
            })

    wb.close()

    daily_df = pd.DataFrame.from_records(daily_records)
    plant_df = pd.DataFrame.from_records(plant_records)
    return daily_df, plant_df


def load_all_months(
    file_path_or_buffer,
    year: Optional[int] = None,
) -> Tuple[pd.DataFrame, pd.DataFrame]:
    """
    Load every month sheet. Returns (daily_df, plant_df) concatenated across months.
    """
    if year is None:
        year = datetime.now().year

    months = list_month_sheets(file_path_or_buffer)
    all_daily = []
    all_plant = []
    for sheet_name, month_num, _display in months:
        d, p = load_month_sheet(file_path_or_buffer, sheet_name, month_num, year=year)
        if not d.empty:
            all_daily.append(d)
        if not p.empty:
            all_plant.append(p)

    if not all_daily:
        return pd.DataFrame(), pd.DataFrame()

    daily = pd.concat(all_daily, ignore_index=True)
    plant = pd.concat(all_plant, ignore_index=True) if all_plant else pd.DataFrame()

    # Deduplicate daily: if the same (vehicle, date) appears in two sheets
    # (e.g. April 1 in both March and April sheets), keep the row whose sheet
    # matches the cell's month.
    daily["_date_month"] = daily["date"].dt.month
    daily["_sheet_month"] = daily["month_name"].str[:3].str.lower().map(
        {k: v[0] for k, v in _MONTH_PATTERNS.items()}
    )
    daily["_match"] = daily["_date_month"] == daily["_sheet_month"]
    daily = daily.sort_values(
        ["vehicle", "date", "_match"], ascending=[True, True, False]
    )
    daily = daily.drop_duplicates(subset=["vehicle", "date"], keep="first")
    daily = daily.drop(columns=["_date_month", "_sheet_month", "_match"])
    daily = daily.reset_index(drop=True)

    return daily, plant
