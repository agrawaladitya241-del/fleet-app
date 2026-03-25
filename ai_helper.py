import openpyxl
import re


def clean_vehicle(v):
    if isinstance(v, str):
        return v.replace(" ", "").upper()
    return ""


def extract_vehicle(text):
    text = text.upper().replace(" ", "")
    match = re.findall(r'[A-Z]{2}[0-9]{2}[A-Z]{2}[0-9]{4}', text)
    return match[0] if match else None


def process_file(uploaded_file):
    wb = openpyxl.load_workbook(uploaded_file)
    ws = wb.active

    vehicle_col = None
    trips_col = None

    for row in ws.iter_rows(min_row=1, max_row=10):
        for i, cell in enumerate(row):
            if cell.value:
                text = str(cell.value).upper()
                if "VEHICLE" in text:
                    vehicle_col = i
                if "TRIP" in text:
                    trips_col = i

    data = {}

    if vehicle_col is None or trips_col is None:
        return data

    for row in ws.iter_rows(min_row=2):

        vehicle_raw = row[vehicle_col].value
        if not vehicle_raw:
            continue

        vehicle = clean_vehicle(vehicle_raw)

        try:
            trips = int(row[trips_col].value)
        except:
            trips = 0

        idle = 0
        for i, cell in enumerate(row):
            if i in [vehicle_col, trips_col]:
                continue

            val = str(cell.value).upper() if cell.value else ""
            if "WAIT" in val or "PARK" in val:
                idle += 1

        if vehicle not in data:
            data[vehicle] = {"trips": 0, "idle": 0}

        data[vehicle]["trips"] += trips
        data[vehicle]["idle"] += idle

    return data


def merge_files(files):
    final_data = {}

    for f in files:
        file_data = process_file(f)

        for v, stats in file_data.items():
            if v not in final_data:
                final_data[v] = {"trips": 0, "idle": 0}

            final_data[v]["trips"] += stats["trips"]
            final_data[v]["idle"] += stats["idle"]

    return final_data


def fleet_summary(files):
    data = merge_files(files)

    total_vehicles = len(data)
    total_trips = sum(v["trips"] for v in data.values())
    total_idle = sum(v["idle"] for v in data.values())

    efficiency = round(total_trips / (total_trips + total_idle), 3) if (total_trips + total_idle) else 0

    return {
        "total_vehicles": total_vehicles,
        "total_trips": total_trips,
        "total_idle": total_idle,
        "efficiency": efficiency,
        "vehicle_data": data
    }


def smart_query(user_input, files):
    summary = fleet_summary(files)
    data = summary["vehicle_data"]

    text = user_input.lower()

    vehicle = extract_vehicle(user_input)
    if vehicle and vehicle in data:
        d = data[vehicle]
        return f"{vehicle} → Trips: {d['trips']}, Idle: {d['idle']}"

    if "total" in text and "trip" in text:
        return f"Total trips: {summary['total_trips']}"

    if "idle" in text:
        return f"Total idle: {summary['total_idle']}"

    if "best" in text:
        top = sorted(data.items(), key=lambda x: x[1]["trips"], reverse=True)[:5]
        return "\n".join([f"{v} → {d['trips']} trips" for v, d in top])

    if "worst" in text:
        worst = sorted(data.items(), key=lambda x: x[1]["trips"])[:5]
        return "\n".join([f"{v} → {d['trips']} trips" for v, d in worst])

    return "Ask about trips, idle, best, worst or vehicle number"


def compare_files(file1, file2):

    data1 = process_file(file1)
    data2 = process_file(file2)

    result = {}

    vehicles = set(data1.keys()).union(set(data2.keys()))

    for v in vehicles:
        d1 = data1.get(v, {"trips": 0, "idle": 0})
        d2 = data2.get(v, {"trips": 0, "idle": 0})

        result[v] = {
            "trip_change": d2["trips"] - d1["trips"],
            "idle_change": d2["idle"] - d1["idle"]
        }

    return result
